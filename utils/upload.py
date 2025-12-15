import os
import logging
import json
import numpy as np
import pandas as pd

import gspread
from dotenv import load_dotenv
from google.oauth2.service_account import Credentials
from openpyxl import load_workbook

from utils.sheet_manager import get_current_sheet_id

load_dotenv()

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

SERVICE_ACCOUNT_JSON_STRING = os.getenv("GOOGLE_CREDENTIALS_JSON")


def sanitize_dataframe(df):
    # First pass: convert datetime columns to strings
    df_copy = df.copy()
    for col in df_copy.columns:
        if df_copy[col].dtype == 'datetime64[ns]':
            df_copy[col] = df_copy[col].astype(str)
    
    # Second pass: fill ALL NaN with empty string
    df_copy = df_copy.fillna('')
    
    # Third pass: replace inf with empty string
    df_copy = df_copy.replace([np.inf, -np.inf], '')
    
    # Fourth pass: convert each value to proper Python types
    result = []
    for _, row in df_copy.iterrows():
        clean_row = []
        for val in row:
            # Handle different types
            if pd.isna(val):
                clean_row.append('')
            elif isinstance(val, (np.integer, np.floating)):
                if np.isnan(val) or np.isinf(val):
                    clean_row.append('')
                else:
                    clean_row.append(float(val))  # Convert to Python float
            elif isinstance(val, pd.Timestamp):
                clean_row.append(str(val))
            elif isinstance(val, (list, tuple, dict, pd.Series)):
                # If it's a complex type, convert to string
                clean_row.append(str(val))
            else:
                # For everything else (strings, etc.)
                clean_row.append(str(val) if val != '' else '')
        result.append(clean_row)
    
    return result


def upload(df, table_name, sheet_name, sheet_id=None):
    """
    Upload the sheet into the given sheet ID and sheet name

    Parameters:
        df (pd.DataFrame): The DataFrame which needs to be uploaded
        table_name (str): The type of table which needs to be uploaded, eg: master_sheet
        sheet_name (str): The name of the worksheet to upload to, eg: Pivot Table
        sheet_id (str, optional): Explicit Google Sheet ID to target. If omitted, falls back to Supabase lookup.
    """
    SHEET_ID = sheet_id or get_current_sheet_id(table_name)

    if not SHEET_ID:
        raise ValueError("Sheet ID is required to upload data.")
    WORKSHEET_TITLE = sheet_name

    logger.info("Starting upload to Google Sheet: %s, worksheet: %s", SHEET_ID, WORKSHEET_TITLE)

    # Authenticate
    scopes = ["https://www.googleapis.com/auth/spreadsheets"]
    logger.info("Authenticating with Google Sheets API")
    service_account_info = json.loads(SERVICE_ACCOUNT_JSON_STRING)
    credentials = Credentials.from_service_account_info(service_account_info, scopes=scopes)
    client = gspread.authorize(credentials)

    logger.info("Opening Google Sheet by key")
    sheet = client.open_by_key(SHEET_ID)

    try:
        logger.info("Trying to add new worksheet: %s", WORKSHEET_TITLE)
        worksheet = sheet.add_worksheet(title=WORKSHEET_TITLE, rows="100", cols="20")
        logger.info("Worksheet created: %s", WORKSHEET_TITLE)

    except Exception as e:
        logger.info("Worksheet already exists. Selecting and clearing worksheet: %s", WORKSHEET_TITLE)
        worksheet = sheet.worksheet(WORKSHEET_TITLE)
        worksheet.clear()

    logger.info("Preparing DataFrame for upload")
    
    # Check if this is a pivot table (case-insensitive check)
    is_pivot = "pivot" in sheet_name.lower()
    
    if is_pivot:
        logger.info("Detected pivot table - using simple conversion")
        
        # Reset index to convert it to a column (NOT drop=True!)
        df_copy = df.reset_index()
        
        # Fill NaN values
        df_copy = df_copy.fillna('')
        
        # Get headers
        headers = df_copy.columns.tolist()
        
        # Convert values to list - simple approach
        data_rows = df_copy.values.tolist()
        
        # Convert numpy types to Python types
        clean_data_rows = []
        for row in data_rows:
            clean_row = []
            for val in row:
                if isinstance(val, (np.integer, np.int64, np.int32)):
                    clean_row.append(int(val))
                elif isinstance(val, (np.floating, np.float64, np.float32)):
                    clean_row.append(float(val))
                elif pd.isna(val) or val == '':
                    clean_row.append('')
                else:
                    clean_row.append(str(val))
            clean_data_rows.append(clean_row)
        
        values = [headers] + clean_data_rows
        
        logger.info(f"Pivot table prepared with {len(values)} rows (including header)")
        logger.info(f"Headers: {headers}")
        logger.info(f"First data row: {clean_data_rows[0] if clean_data_rows else 'No data'}")
        
    else:
        logger.info("Regular table - applying full sanitization")
        # Get headers from the original DataFrame
        headers = df.columns.tolist()
        
        # Get sanitized values (returns a list of lists)
        clean_values = sanitize_dataframe(df)
        
        # Combine headers and body
        values = [headers] + clean_values

    logger.info("Uploading data to worksheet")
    worksheet.update("A1", values)
    logger.info("Upload complete")


def append_rolling_data(df, table_name, sheet_name, max_rows=300000):
    """
    Appends data to a sheet, ensuring the total row count does not exceed max_rows.
    Deletes oldest rows BEFORE appending to avoid exceeding Google Sheets' cell limit.
    
    NOTE: 
    - max_rows includes the header row
    - Default max_rows set to 300,000 to avoid hitting Google Sheets' 10M cell limit per workbook
    - For a 26-column sheet: 300k rows × 26 cols = 7.8M cells (safe margin for workbook)
    - Row 1 is always the header and is never deleted
    - Uses gspread delete_rows which has INCLUSIVE start and end indices
    - IMPORTANT: Deletes old data BEFORE appending to avoid cell limit errors during append
    """
    # Validate input
    if df is None or df.empty:
        logger.warning("Empty DataFrame provided to append_rolling_data. Skipping append.")
        return
    
    SHEET_ID = get_current_sheet_id(table_name)
    
    # Authenticate
    scopes = ["https://www.googleapis.com/auth/spreadsheets"]
    
    service_account_info = json.loads(SERVICE_ACCOUNT_JSON_STRING)
    credentials = Credentials.from_service_account_info(service_account_info, scopes=scopes)
    
    client = gspread.authorize(credentials)
    
    logger.info("Opening Google Sheet for Rolling Append: %s", SHEET_ID)
    sheet = client.open_by_key(SHEET_ID)

    # Get or create worksheet
    try:
        worksheet = sheet.worksheet(sheet_name)
        logger.info("Worksheet '%s' found", sheet_name)
    except gspread.WorksheetNotFound:
        logger.info("Worksheet '%s' not found. Creating new.", sheet_name)
        worksheet = sheet.add_worksheet(title=sheet_name, rows="100", cols="20")
        
        # Initialize new worksheet with headers and data
        headers = [df.columns.tolist()]
        data_rows = sanitize_dataframe(df)
        values = headers + data_rows
        
        worksheet.update("A1", values)
        logger.info("New worksheet created and populated with %d rows (including header)", len(values))
        return

    # Get current data count
    current_data = worksheet.get_all_values()
    current_row_count = len(current_data)  # Includes header
    new_row_count = len(df)  # Data rows only, no header
    
    logger.info("Current rows (with header): %d, New rows: %d, Limit: %d", 
                current_row_count, new_row_count, max_rows)

    # Calculate what the final row count should be after append
    final_row_count = min(current_row_count + new_row_count, max_rows)
    
    # Calculate how many rows to delete BEFORE appending
    # We want to make room for new data while staying under max_rows
    if current_row_count + new_row_count > max_rows:
        rows_to_delete = current_row_count + new_row_count - max_rows
        logger.info("Will exceed limit. Deleting %d oldest data rows BEFORE appending...", rows_to_delete)
        
        # Calculate max deletable rows (can't delete header at row 1)
        max_deletable = current_row_count - 1  # -1 for header
        
        if rows_to_delete > max_deletable:
            logger.warning("Need to delete %d rows but only %d data rows exist. Deleting all data rows.", 
                         rows_to_delete, max_deletable)
            rows_to_delete = max_deletable
        
        if rows_to_delete > 0:
            start_index = 2
            end_index = 2 + rows_to_delete - 1  # Inclusive end
            
            logger.info("Deleting rows %d to %d (%d rows) to make room...", start_index, end_index, rows_to_delete)
            worksheet.delete_rows(start_index, end_index)
            logger.info("Deletion complete. Sheet now has %d rows (including header)", current_row_count - rows_to_delete)

    # Now append new data (we've made room)
    logger.info("Appending %d new rows...", new_row_count)
    data_to_add = sanitize_dataframe(df)
    
    worksheet.append_rows(data_to_add, value_input_option="USER_ENTERED")
    logger.info("Append complete. Added %d rows. Final count should be ~%d rows.", len(data_to_add), final_row_count)


def upload_excel(excel_path: str, table_name: str, sheet_name: str, sheet_id: str = None):
    """
    Upload an Excel file directly to a Google Sheet without using Drive import.
    Converts the Excel into values and applies basic formatting (bold/fills) based
    on the Excel styles.

    Parameters:
        excel_path (str): Path to the Excel file to upload.
        table_name (str): Table name used to look up the default Sheet ID.
        sheet_name (str): Target worksheet title inside the destination spreadsheet.
        sheet_id (str, optional): Explicit Google Sheet ID to target. If omitted, falls back to Supabase lookup.
    """
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    SHEET_ID = sheet_id or get_current_sheet_id(table_name)
    if not SHEET_ID:
        raise ValueError("Sheet ID is required to upload data.")

    logger.info("Starting Excel upload to Google Sheet: %s, worksheet: %s", SHEET_ID, sheet_name)

    # Read Excel values and basic styles
    wb = load_workbook(excel_path)
    ws = wb.active
    max_row, max_col = ws.max_row, ws.max_column

    values = []
    bold_cells = {}
    header_fill_row = None
    header_fill_color = None

    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col), start=1):
        row_values = []
        for c_idx, cell in enumerate(row, start=1):
            row_values.append("" if cell.value is None else cell.value)

            if cell.font and cell.font.bold:
                bold_cells.setdefault(r_idx, []).append(c_idx)

            if (
                cell.fill
                and cell.fill.fill_type == "solid"
                and cell.fill.start_color
                and cell.fill.start_color.rgb
                and cell.fill.start_color.rgb.lower() not in ("00000000", "ffffffff")
            ):
                header_fill_row = r_idx
                header_fill_color = cell.fill.start_color.rgb

        values.append(row_values)

    # Authenticate (Sheets scope only)
    scopes = ["https://www.googleapis.com/auth/spreadsheets"]
    service_account_info = json.loads(SERVICE_ACCOUNT_JSON_STRING)
    credentials = Credentials.from_service_account_info(service_account_info, scopes=scopes)
    client = gspread.authorize(credentials)
    sheet = client.open_by_key(SHEET_ID)

    # Delete existing worksheet with the same name (if present) to avoid conflicts.
    try:
        existing_ws = sheet.worksheet(sheet_name)
        sheet.del_worksheet(existing_ws)
        logger.info("Deleted existing worksheet named '%s' before upload", sheet_name)
    except gspread.WorksheetNotFound:
        logger.info("No existing worksheet named '%s' found; continuing", sheet_name)

    # Create new worksheet sized to the data
    worksheet = sheet.add_worksheet(
        title=sheet_name,
        rows=str(max(len(values), 100)),
        cols=str(max_col + 5),
    )

    worksheet_sheet_id = getattr(worksheet, "id", None) or worksheet._properties.get("sheetId")

    # Upload values
    worksheet.update("A1", values)

    # Build formatting requests (bold, header fill/alignment, column widths)
    requests = []

    # Helper to convert ARGB hex to Sheets color dict
    def _hex_to_color(argb: str) -> dict:
        hex_part = argb[-6:]  # drop alpha if present
        r = int(hex_part[0:2], 16) / 255
        g = int(hex_part[2:4], 16) / 255
        b = int(hex_part[4:6], 16) / 255
        return {"red": r, "green": g, "blue": b}

    # Header fill + alignment (apply to the row that had a fill in Excel)
    if header_fill_row and header_fill_color:
        requests.append(
            {
                "repeatCell": {
                    "range": {
                        "sheetId": worksheet_sheet_id,
                        "startRowIndex": header_fill_row - 1,
                        "endRowIndex": header_fill_row,
                        "startColumnIndex": 0,
                        "endColumnIndex": max_col,
                    },
                    "cell": {
                        "userEnteredFormat": {
                            "backgroundColor": _hex_to_color(header_fill_color),
                            "horizontalAlignment": "CENTER",
                            "verticalAlignment": "MIDDLE",
                            "textFormat": {"bold": True},
                        }
                    },
                    "fields": "userEnteredFormat(backgroundColor,horizontalAlignment,verticalAlignment,textFormat.bold)",
                }
            }
        )

    # Bold cells (apply only to cells that were bold in Excel)
    for r_idx, cols in bold_cells.items():
        cols = sorted(cols)
        start = cols[0]
        end = cols[0]
        for col in cols[1:]:
            if col == end + 1:
                end = col
            else:
                requests.append(
                    {
                        "repeatCell": {
                            "range": {
                                "sheetId": worksheet_sheet_id,
                                "startRowIndex": r_idx - 1,
                                "endRowIndex": r_idx,
                                "startColumnIndex": start - 1,
                                "endColumnIndex": end,
                            },
                            "cell": {"userEnteredFormat": {"textFormat": {"bold": True}}},
                            "fields": "userEnteredFormat.textFormat.bold",
                        }
                    }
                )
                start = col
                end = col
        requests.append(
            {
                "repeatCell": {
                    "range": {
                        "sheetId": worksheet_sheet_id,
                        "startRowIndex": r_idx - 1,
                        "endRowIndex": r_idx,
                        "startColumnIndex": start - 1,
                        "endColumnIndex": end,
                    },
                    "cell": {"userEnteredFormat": {"textFormat": {"bold": True}}},
                    "fields": "userEnteredFormat.textFormat.bold",
                }
            }
        )

    # Column widths: approximate based on content length (cap at 50 chars)
    col_widths = []
    for col_idx in range(max_col):
        max_len = 0
        for row in values:
            if col_idx < len(row):
                val_len = len(str(row[col_idx])) if row[col_idx] is not None else 0
                if val_len > max_len:
                    max_len = val_len
        col_widths.append(min(max_len + 2, 50))

    for idx, width in enumerate(col_widths):
        # Rough conversion: 1 char ~ 7 pixels
        pixel_size = int(width * 7)
        requests.append(
            {
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": worksheet_sheet_id,
                        "dimension": "COLUMNS",
                        "startIndex": idx,
                        "endIndex": idx + 1,
                    },
                    "properties": {"pixelSize": pixel_size},
                    "fields": "pixelSize",
                }
            }
        )

    if requests:
        sheet.batch_update({"requests": requests})
        logger.info("Applied %d formatting requests", len(requests))
