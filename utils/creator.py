import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from utils.db import fetch


def get_company_config(configs, company_name):
    """Get configuration for a specific company."""
    for config in configs:
        if config['company_name'] == company_name:
            return config
    raise ValueError(f"Configuration for company '{company_name}' not found")


def create_excel_from_breakdown(breakdown_csv_path, config, output_excel_path):
    """
    Create an Excel file from breakdown CSV using the provided config.
    
    Parameters:
    - breakdown_csv_path: Path to the breakdown CSV file
    - config: Configuration dictionary for the company
    - output_excel_path: Path where the output Excel file will be saved
    """
    # Read the breakdown CSV
    df_breakdown = pd.read_csv(breakdown_csv_path)
    
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    
    current_row = 1
    
    # Add headers from config
    for header_text in config['headers']:
        ws.cell(row=current_row, column=1, value=header_text)
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
        current_row += 1
    
    # Add line gaps
    current_row += config['line_gaps']
    
    # Create the new dataframe with remapped columns
    column_mapping = config['column_mapping']
    new_data = {}
    
    # Map columns from breakdown to new column names
    for new_col, old_col in column_mapping.items():
        if old_col in df_breakdown.columns:
            new_data[new_col] = df_breakdown[old_col]
        else:
            print(f"Warning: Column '{old_col}' not found in breakdown file")
            new_data[new_col] = [''] * len(df_breakdown)
    
    # Add empty UTR column
    utr_column = config.get('utr_column_name', 'UTR')
    new_data[utr_column] = [''] * len(df_breakdown)
    
    # Create new dataframe
    df_new = pd.DataFrame(new_data)
    
    # Write column headers
    header_row = current_row
    for col_idx, column_name in enumerate(df_new.columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=column_name)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    current_row += 1
    
    # Write data rows
    for row_idx, row_data in enumerate(df_new.values, start=current_row):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook
    wb.save(output_excel_path)
    print(f"Excel file created successfully: {output_excel_path}")
    return output_excel_path


def process_breakdown(breakdown_csv_path, company_name, output_excel_path=None):
    """
    Main function to process a breakdown file for a specific company.
    
    Parameters:
    - breakdown_csv_path: Path to the breakdown CSV file
    - config_file_path: Path to the config JSON file
    - company_name: Name of the company to process
    - output_excel_path: Optional output path (default: company_name_output.xlsx)
    """
    configs = fetch("configs/utr_config.json")
    
    # Get the specific company config
    config = get_company_config(configs, company_name)
    
    # Set default output path if not provided
    if output_excel_path is None:
        output_excel_path = f"{company_name}_output.xlsx"
    
    # Create the Excel file
    create_excel_from_breakdown(breakdown_csv_path, config, output_excel_path)
    
    return output_excel_path


# Example usage
if __name__ == "__main__":
    # Example: Process OLAMBIT breakdown
    breakdown_file = "temp/breakdown.csv"
    company = "PURPLE_KUIK"
    
    output_file = process_breakdown(
        breakdown_csv_path=breakdown_file,
        company_name=company,
        output_excel_path="formatted.xlsx"
    )
    
    print(f"Processing complete! Output saved to: {output_file}")